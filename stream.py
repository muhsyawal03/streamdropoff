from flask import Flask, render_template, Response, request
from flask_socketio import SocketIO
import cv2
import time
import numpy as np
from ultralytics import YOLO
import requests
import subprocess
import threading
from collections import OrderedDict
import warnings
import json
from datetime import datetime
import os
from openpyxl import Workbook, load_workbook

os.makedirs('static/pelanggaran', exist_ok=True)


app = Flask(__name__, static_folder='static')
socketio = SocketIO(app, cors_allowed_origins="*")

model = YOLO("yolo/yolo11n.pt")

PARKING_AREA = np.array([
    (20, 700),
    (1250, 700),
    (1100, 220),
    (150, 220)
], np.int32)

VALID_CLASSES = [2, 3]
THRESHOLD_TIME = 10
DISAPPEAR_LIMIT = 10

vehicle_tracker = OrderedDict()
next_vehicle_id = {
    "car": 1,
    "motorcycle": 1
}

TELEGRAM_TOKEN = "8015009177:AAFs3-khMeu0hVkqYPFvKEHSFIC4DH6Za-s"
TELEGRAM_CHAT_ID = "7673995490"
telegram_response_status = {}

warnings.filterwarnings("ignore", category=UserWarning, module='cv2')

# Di global scope, ganti telegram_response_status dengan satu counter saja:
telegram_counter = 0
telegram_done = 0

def send_telegram(message, vehicle_id):
    global telegram_counter

    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
    keyboard = {
        "inline_keyboard": [[
            {"text": "✅ Done", "callback_data": f"done_{vehicle_id}"}
        ]]
    }
    data = {
        "chat_id": TELEGRAM_CHAT_ID,
        "text": message,
        "reply_markup": json.dumps(keyboard)
    }

    try:
        resp = requests.post(url, data=data)
        if resp.status_code != 200:
            print("Telegram Error:", resp.text)
            return

        # setiap notifikasi terkirim, counter +1
        telegram_counter += 1
        # kirim ke frontend
        socketio.emit('update_stats', {
            "respon": telegram_counter,    # gunakan sebagai “pelanggaran” di UI
            "pelanggaran": telegram_counter  # atau jika mau dibalik namanya
        })

    except Exception as e:
        print("Gagal kirim Telegram:", e)


def is_inside_parking(x1, y1, x2, y2):
    cx, cy = (x1 + x2) // 2, (y1 + y2) // 2
    return cv2.pointPolygonTest(PARKING_AREA, (cx, cy), False) >= 0


def euclidean_distance(p1, p2):
    return np.linalg.norm(np.array(p1) - np.array(p2))


def assign_vehicle_id(centroid, cls):
    global next_vehicle_id

    prefix = "M" if cls == 2 else "B"  

    for vid, data in vehicle_tracker.items():
        if euclidean_distance(centroid, data["centroid"]) < 50 and data["class"] == cls:
            data["centroid"] = centroid
            data["frames_absent"] = 0
            return vid

    # Buat ID unik dengan prefix jenis kendaraan
    id_key = f"{prefix}{next_vehicle_id['car' if cls == 2 else 'motorcycle']}"
    vehicle_tracker[id_key] = {
        "centroid": centroid,
        "frames_absent": 0,
        "time_entered": time.time(),
        "notified": False,
        "class": cls
    }

    if cls == 2:
        next_vehicle_id["car"] += 1
    else:
        next_vehicle_id["motorcycle"] += 1

    return id_key


def handle_countdown(vehicle_id):
    countdown_time = 100
    for i in range(countdown_time, 0, -1):
        # Jika kendaraan sudah merespon
        if telegram_response_status.get(vehicle_id) == 0:
            return

        # Jika kendaraan sudah tidak terdeteksi (sudah meninggalkan area)
        if vehicle_id not in vehicle_tracker or vehicle_tracker[vehicle_id]["frames_absent"] > DISAPPEAR_LIMIT:
            msg = f"✅ Kendaraan ID {vehicle_id} sudah meninggalkan area sebelum countdown selesai."
            print(f"[INFO] {msg}")
            socketio.emit('notification', {'message': msg, 'alert': False})

            # Reset status Telegram dan pelanggaran
            telegram_response_status.pop(vehicle_id, None)  # Reset status Telegram
            socketio.emit('telegram_status', {'vehicle_id': vehicle_id, 'status': 0})  # Reset status Telegram di UI

            # Reset penghitung pelanggaran dan respon ke 0
            global telegram_counter, telegram_done
            telegram_counter = 0
            telegram_done = 0

            # Emit reset status pelanggaran dan respon
            socketio.emit('update_stats', {
                "respon": telegram_done,  # Reset statistik respon ke 0
                "pelanggaran": telegram_counter  # Reset statistik pelanggaran ke 0
            })

            return

        socketio.emit('countdown_update', {'vehicle_id': vehicle_id, 'time_left': i})
        time.sleep(1)

    # Countdown habis, kirim notifikasi pelanggaran ulang
    if telegram_response_status.get(vehicle_id) != 0:
        msg = f"🚨 Pelanggaran ulang: Kendaraan ID {vehicle_id} tidak merespon dalam 100 detik!"
        socketio.emit('notification', {'message': msg, 'alert': True})
        send_telegram(msg, vehicle_id)

        # Reset penghitung pelanggaran dan respon ke 1
        telegram_counter += 1  # Increment pelanggaran setelah countdown habis
        telegram_done = 1  # Menghitung respon untuk pelanggaran pertama
        socketio.emit('update_stats', {"pelanggaran": telegram_counter, "respon": telegram_done})
        socketio.emit('telegram_status', {'vehicle_id': vehicle_id, 'status': 1})

def generate_frames():
    cap = cv2.VideoCapture('http://192.168.101.149:8080/?action=stream')
    if not cap.isOpened():
        while True:
            frame = np.zeros((480, 640, 3), dtype=np.uint8)
            cv2.putText(frame, "No Camera Detected", (150, 240),
                        cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)
            _, buffer = cv2.imencode('.jpg', frame)
            yield (b'--frame\r\nContent-Typ e: image/jpeg\r\n\r\n' + buffer.tobytes())
        return
    global vehicle_tracker
    while True:
        now = datetime.now()
        current_hour = now.hour
        if current_hour < 7 or current_hour >= 17:
            frame = np.zeros((480, 640, 3), dtype=np.uint8)
            msg = "CCTV tidak aktif di luar jam 07:00 - 17:00"
            cv2.putText(frame, msg, (50, 240), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 0, 255), 2)
            _, buffer = cv2.imencode('.jpg', frame)
            yield (b'--frame\r\nContent-Type: image/jpeg\r\n\r\n' + buffer.tobytes())
            time.sleep(1)
            continue
        success, frame = cap.read()
        if not success:
            continue
        results = model(frame)
        detected_vehicles = set()
        confidences = []
        for result in results:
            for box in result.boxes:
                cls = int(box.cls[0])
                if cls in VALID_CLASSES:
                    x1, y1, x2, y2 = map(int, box.xyxy[0])
                    conf = float(box.conf[0]) * 100
                    confidences.append(conf)
                    centroid = ((x1 + x2) // 2, (y1 + y2) // 2)
                    vehicle_id = assign_vehicle_id(centroid, cls)
                    detected_vehicles.add(vehicle_id)
                    
                    box_color = (0, 255, 0)  # Hijau default
                    cv2.rectangle(frame, (x1, y1), (x2, y2), box_color, 2)
                    cv2.putText(frame, f"ID {vehicle_id}", (x1, y1 - 10),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 255), 2)

                    if is_inside_parking(x1, y1, x2, y2):
                        tracked = vehicle_tracker[vehicle_id]
                        parked_time = time.time() - tracked["time_entered"]
                       
                        if parked_time > THRESHOLD_TIME:
                            box_color = (0, 0, 255)  # Merah jika melanggar
                            cv2.rectangle(frame, (x1, y1), (x2, y2), box_color, 2)

                            # Jika belum ada yang dilaporkan, lakukan proses pelanggaran kolektif
                            if not any([v.get("notified") for v in vehicle_tracker.values()]):
                                msg = f"🚨 Pelanggaran: Kendaraan ID {vehicle_id} parkir lebih dari {THRESHOLD_TIME} detik!"
                                socketio.emit('notification', {'message': msg, 'alert': True})
                                send_telegram(msg, "semua")

                                for vid in vehicle_tracker:
                                    threading.Thread(target=handle_countdown, args=(vid,)).start()
                                    vehicle_tracker[vid]["notified"] = True

                                # === Info Umum Pelanggaran ===
                                now = datetime.now()
                                tanggal = now.strftime('%d %B %Y')  # ex: 30 Juni 2025
                                waktu = now.strftime('%H:%M:%S')
                                hari = now.strftime('%A')            # ex: Sunday
                                pelanggaran_time = now.strftime('%Y%m%d_%H%M%S')
                                filename = f"pelanggaran_{pelanggaran_time}.jpg"
                                filepath = f"static/pelanggaran/{filename}"
                                os.makedirs("static/pelanggaran", exist_ok=True)

                                # === Kumpulkan Info Pelanggar ===
                                text_lines = [f"{hari}, {tanggal} {waktu}"]  # header tanggal
                                excel_rows = []

                                for vid, data in vehicle_tracker.items():
                                    vehicle_class = "mobil" if data["class"] == 2 else "motor"
                                    avg_conf = sum(confidences) / len(confidences) if confidences else 0.0
                                    text_lines.append(f"ID: {vid} | Akurasi: {avg_conf:.2f}%")
                                    excel_rows.append([tanggal, hari, waktu, vid, vehicle_class, f"{avg_conf:.2f}", filename])

                                # === Tampilkan Info di Gambar ===
                                overlay = frame.copy()
                                start_x = 10
                                start_y = 40
                                line_height = 30
                                max_width = 520
                                box_height = line_height * len(text_lines) + 20

                                cv2.rectangle(overlay, (start_x - 5, start_y - 25),
                                            (start_x + max_width, start_y - 25 + box_height),
                                            (0, 0, 0), -1)
                                cv2.addWeighted(overlay, 0.5, frame, 0.5, 0, frame)

                                for i, line in enumerate(text_lines):
                                    cv2.putText(frame, line, (start_x, start_y + i * line_height),
                                                cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)

                                # === Simpan Gambar ===
                                os.makedirs("static/pelanggaran", exist_ok=True)
                                cv2.imwrite(filepath, frame)

                                # === Simpan ke Excel ===
                                log_file = 'static/pelanggaran_log.xlsx'
                                headers = ['tanggal', 'hari', 'waktu', 'vehicle_id', 'kelas', 'confidence', 'nama_file']

                                if not os.path.exists(log_file):
                                    wb = Workbook()
                                    ws = wb.active
                                    ws.append(headers)
                                else:
                                    wb = load_workbook(log_file)
                                    ws = wb.active

                                for row in excel_rows:
                                    ws.append(row)

                                wb.save(log_file)

        # Hapus kendaraan yang tidak terdeteksi lagi
        for vid in list(vehicle_tracker.keys()):
            if vid not in detected_vehicles:
                vehicle_tracker[vid]["frames_absent"] += 1
                if vehicle_tracker[vid]["frames_absent"] > DISAPPEAR_LIMIT:
                    del vehicle_tracker[vid]

        # Gambar area parkir
        cv2.polylines(frame, [PARKING_AREA], isClosed=True, color=(0, 255, 255), thickness=2)

        # Tambahkan timestamp
        timestamp = time.strftime('%Y-%m-%d %H:%M:%S')
        cv2.putText(frame, timestamp, (10, frame.shape[0] - 10),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.6, (180, 180, 180), 1)

        # Tambahkan akurasi
        if confidences:
            avg_conf = sum(confidences) / len(confidences)
            overlay = frame.copy()
            cv2.rectangle(overlay, (frame.shape[1] - 230, 10), (frame.shape[1] - 10, 45), (0, 0, 0), -1)
            alpha = 0.4
            cv2.addWeighted(overlay, alpha, frame, 1 - alpha, 0, frame)
            cv2.putText(frame, f"Akurasi: {avg_conf:.2f}%", (frame.shape[1] - 220, 35),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)

        _, buffer = cv2.imencode('.jpg', frame)
        yield (b'--frame\r\nContent-Type: image/jpeg\r\n\r\n' + buffer.tobytes())

    cap.release()
 

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/video_feed')
def video_feed():
    return Response(generate_frames(), mimetype='multipart/x-mixed-replace; boundary=frame')


@app.route('/test_emit')
def test_emit():
    socketio.emit('notification', {'message': '🔔 Ini tes notifikasi!', 'alert': True})
    return "Emit sent"


@app.route('/set_audio_output', methods=['POST'])
def set_audio_output():
    data = request.get_json()
    device = data.get('device')
    try:
        subprocess.run(['wpctl', 'set-default', device], check=True)
        return "Output diubah", 200
    except Exception as e:
        return f"Error: {str(e)}", 500


@app.route('/telegram_webhook', methods=['POST'])
def telegram_webhook():
    global telegram_counter, telegram_done
    data = request.get_json()

    if 'callback_query' in data and data['callback_query']['data'].startswith('done_'):
        vid = int(data['callback_query']['data'].split('_')[1])

        # Kurangi respon dan pelanggaran
        telegram_counter = max(0, telegram_counter - 1)
        telegram_done = max(0, telegram_done - 1)

        # Edit pesan di Telegram
        chat = data['callback_query']['message']['chat']['id']
        mid = data['callback_query']['message']['message_id']
        requests.post(
            f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/editMessageText",
            data={
                "chat_id": chat,
                "message_id": mid,
                "text": f"✅ Kendaraan ID {vid} telah dikonfirmasi selesai oleh petugas."
            }
        )

        # Hapus data kendaraan
        vehicle_tracker.pop(vid, None)
        telegram_response_status.pop(vid, None)

        # Reset total jika kosong
        if not vehicle_tracker:
            telegram_counter = 0
            telegram_done = 0

        # Emit ke UI setelah semua update
        socketio.emit('update_stats', {
            "respon": telegram_done,
            "pelanggaran": telegram_counter
        })
        socketio.emit('cancel_countdown', {'vehicle_id': vid})

        print(f"[WEBHOOK] Done ditekan untuk kendaraan ID {vid}")
        print(f"[SOCKET] Update stats: Respon={telegram_done}, Pelanggaran={telegram_counter}")
    
    return "OK", 200


@app.route('/status')
def status():
    return telegram_response_status


@app.route('/reset_all', methods=['POST'])
def reset_all():
    vehicle_tracker.clear()
    telegram_response_status.clear()
    socketio.emit('update_stats', {"respon": 0, "pelanggaran": 0})
    return "Semua data telah direset", 200

@app.route('/history')
def history():
    log_file = 'static/pelanggaran_log.xlsx'
    data = []

    if os.path.exists(log_file):
        wb = load_workbook(log_file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            data.append(row)

    return render_template('history.html', data=data)


@socketio.on("play_mic")
def handle_play_mic():
    def play_sound():
        try:
            subprocess.run([
                "C:\\ffmpeg\\bin\\ffplay.exe", "-nodisp", "-autoexit",
                "D:\\TUGAS & PROJECT\\Sem 6\Vision\\project\\static\\sound\\violation_alert.mp3"
            ], check=True)
        except Exception as e:
            print(f"[ERROR] Failed to play audio: {e}")

    threading.Thread(target=play_sound).start()


if __name__ == '__main__':
    socketio.run(app, host='0.0.0.0', port=5000)
